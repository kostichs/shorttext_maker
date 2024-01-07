import openpyxl
import re

# Открываем файл Excel
workbook = openpyxl.load_workbook('Kurztext.xlsx')

# Выбираем активный лист (предположим, что он единственный)
sheet = workbook.active

# Ваши столбцы
columns = {
    "laenge": 28, "breite": 29, "hoehe": 30, "gewicht": 32, "gewichtKG": 31,
    "material": 51, "freiMaterial": 100, "farbe": 35, "freiFarbe": 98,
    "groesse": 36, "freiGroesse": 97, "tauMin": 66, "tauMax": 67,
    "lagerart": 72, "buegel": 77, "unterbuegel": 78, "merkmal": 102,
    "pin": 92, "augbreite": 93, "aughoehe": 94, "elleistung": 56,
    "aussendurchmesser": 41, "kapazitat": 83, "schaftlange": 85,
    "einsatzZweck": 69
}

# Проходим по каждой строке
for row_number in range(2, sheet.max_row + 1):  # Начинаем с 2 строки, предполагая, что первая строка - заголовки
    kurztext_values = []

    # Проверяем каждый столбец
    for column_name, column_index in columns.items():
        cell_value = sheet.cell(row=row_number, column=column_index).value
        if cell_value is not None:
            # Извлекаем единицу измерения из первой строки столбца
            # unit_value = sheet.cell(row=1, column=column_index).value
            title = sheet.cell(row=1, column=column_index).value
            match = re.search(r'\(([^)]+)\)', title)

            if match:
                result = match.group(1)
                if '*' in result:
                    result = result.replace('*', '')
                kurztext_values.append(f"{cell_value} {result}")
            else:
                kurztext_values.append(f"{cell_value}")

    # Записываем значения в ячейку kurztext
    kurztext_cell = sheet.cell(row=row_number, column=27)  # столбец AA
    kurztext_cell.value = "|".join(kurztext_values[:5])

# Сохраняем изменения
workbook.save('новый_файл.xlsx')
